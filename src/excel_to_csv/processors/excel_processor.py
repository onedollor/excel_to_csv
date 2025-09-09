"""Excel file processing for Excel-to-CSV converter.

This module provides robust Excel file processing capabilities including:
- Reading various Excel formats (.xlsx, .xls)
- Extracting worksheet metadata and data
- Error handling for corrupt or locked files
- Memory-efficient processing of large files
"""

import time
from pathlib import Path
from typing import List, Optional, Dict, Any, Union

import pandas as pd
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

from excel_to_csv.models.data_models import WorksheetData
from excel_to_csv.utils.logger import get_processing_logger
from excel_to_csv.utils.logging_decorators import log_operation, operation_context
from excel_to_csv.utils.correlation import CorrelationContext


class ExcelProcessingError(Exception):
    """Raised when Excel file processing fails."""
    pass


class ExcelProcessor:
    """Processes Excel files and extracts worksheet data.
    
    The ExcelProcessor handles:
    - Reading Excel files in various formats
    - Extracting worksheet metadata and data
    - Error handling for various file conditions
    - Memory-efficient processing
    
    Example:
        >>> processor = ExcelProcessor()
        >>> worksheets = processor.process_file("data.xlsx")
        >>> for ws in worksheets:
        ...     print(f"Worksheet: {ws.worksheet_name}, Rows: {ws.row_count}")
    """
    
    # Supported Excel file extensions
    SUPPORTED_EXTENSIONS = {'.xlsx', '.xls'}
    
    # Maximum number of rows to preview for metadata
    METADATA_PREVIEW_ROWS = 100
    
    def __init__(self, max_file_size_mb: int = 100, chunk_size: int = 10000):
        """Initialize Excel processor.
        
        Args:
            max_file_size_mb: Maximum file size to process in MB
            chunk_size: Chunk size for reading large files
        """
        self.max_file_size_mb = max_file_size_mb
        self.chunk_size = chunk_size
        self.logger = get_processing_logger(__name__)
    
    @log_operation("process_excel_file")
    def process_file(self, file_path: Union[str, Path]) -> List[WorksheetData]:
        """Process Excel file and return worksheet data.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            List of WorksheetData objects for each worksheet
            
        Raises:
            ExcelProcessingError: If file cannot be processed
        """
        file_path = Path(file_path)
        
        with operation_context(
            "excel_file_processing",
            self.logger,
            file_path=str(file_path),
            file_size=file_path.stat().st_size if file_path.exists() else 0
        ) as metrics:
            
            # Validate file
            with operation_context("file_validation", self.logger, file_path=str(file_path)):
                self._validate_file(file_path)
            
            # Log processing start
            file_size = file_path.stat().st_size
            file_extension = file_path.suffix.lower()
            
            self.logger.info(
                f"Starting Excel file processing: {file_path} ({file_size:,} bytes)",
                extra={
                    "structured": {
                        "operation": "excel_processing_start", 
                        "file_path": str(file_path),
                        "file_size": file_size,
                        "file_extension": file_extension,
                        "max_file_size": self.max_file_size_mb * 1024 * 1024
                    }
                }
            )
            
            metrics.add_metadata("file_extension", file_extension)
            metrics.add_metadata("file_size_mb", file_size / (1024 * 1024))
            
            try:
                # Extract worksheets
                worksheets = self._extract_worksheets(file_path)
                
                worksheet_count = len(worksheets)
                metrics.add_metadata("worksheets_extracted", worksheet_count)
                
                # Log completion with detailed summary
                total_rows = sum(ws.row_count for ws in worksheets)
                total_columns = sum(ws.column_count for ws in worksheets)
                
                self.logger.info(
                    f"Excel processing completed: {file_path} - {worksheet_count} worksheets, "
                    f"{total_rows:,} total rows, {total_columns} total columns",
                    extra={
                        "structured": {
                            "operation": "excel_processing_complete",
                            "file_path": str(file_path),
                            "worksheets_count": worksheet_count,
                            "total_rows": total_rows,
                            "total_columns": total_columns,
                            "worksheets": [
                                {
                                    "name": ws.worksheet_name,
                                    "rows": ws.row_count,
                                    "columns": ws.column_count,
                                    "has_headers": ws.has_headers,
                                    "data_type_summary": ws.column_types
                                } for ws in worksheets
                            ]
                        }
                    }
                )
                
                metrics.add_metadata("total_rows", total_rows)
                metrics.add_metadata("total_columns", total_columns)
                metrics.add_metadata("processing_rate_rows_per_sec", 
                                   total_rows / metrics.duration_ms * 1000 if metrics.duration_ms else 0)
                
                return worksheets
                
            except Exception as e:
                error_type = type(e).__name__
                metrics.add_metadata("error_type", error_type)
                metrics.add_metadata("error_message", str(e))
                
                self.logger.error(
                    f"Excel processing failed for {file_path}: {e}",
                    exc_info=True,
                    extra={
                        "structured": {
                            "operation": "excel_processing_failed",
                            "file_path": str(file_path),
                            "error_type": error_type,
                            "error_message": str(e)
                        }
                    }
                )
                raise ExcelProcessingError(f"Excel processing failed: {e}") from e
    
    @log_operation("validate_excel_file", log_args=False)
    def _validate_file(self, file_path: Path) -> None:
        """Validate Excel file before processing.
        
        Args:
            file_path: Path to Excel file
            
        Raises:
            ExcelProcessingError: If validation fails
        """
        with operation_context(
            "file_validation",
            self.logger,
            file_path=str(file_path)
        ) as metrics:
            
            self.logger.info(f"Validating Excel file: {file_path}")
            
            # Check if file exists
            if not file_path.exists():
                error_msg = f"File not found: {file_path}"
                metrics.add_metadata("validation_error", "file_not_found")
                self.logger.error(error_msg, extra={"structured": {"operation": "file_validation_failed", "error_type": "file_not_found", "file_path": str(file_path)}})
                raise ExcelProcessingError(error_msg)
            
            # Check if it's a file (not directory)
            if not file_path.is_file():
                error_msg = f"Path is not a file: {file_path}"
                metrics.add_metadata("validation_error", "not_a_file")
                self.logger.error(error_msg, extra={"structured": {"operation": "file_validation_failed", "error_type": "not_a_file", "file_path": str(file_path)}})
                raise ExcelProcessingError(error_msg)
            
            # Check file extension
            if file_path.suffix.lower() not in self.SUPPORTED_EXTENSIONS:
                error_msg = f"Unsupported file extension: {file_path.suffix}. Supported: {', '.join(self.SUPPORTED_EXTENSIONS)}"
                metrics.add_metadata("validation_error", "unsupported_extension")
                metrics.add_metadata("file_extension", file_path.suffix.lower())
                self.logger.error(error_msg, extra={"structured": {"operation": "file_validation_failed", "error_type": "unsupported_extension", "file_path": str(file_path), "extension": file_path.suffix.lower()}})
                raise ExcelProcessingError(error_msg)
            
            # Check file size
            file_size = file_path.stat().st_size
            file_size_mb = file_size / (1024 * 1024)
            metrics.add_metadata("file_size_bytes", file_size)
            metrics.add_metadata("file_size_mb", file_size_mb)
            
            if file_size_mb > self.max_file_size_mb:
                error_msg = f"File too large: {file_size_mb:.1f}MB > {self.max_file_size_mb}MB"
                metrics.add_metadata("validation_error", "file_too_large")
                self.logger.error(error_msg, extra={"structured": {"operation": "file_validation_failed", "error_type": "file_too_large", "file_path": str(file_path), "file_size_mb": file_size_mb, "max_size_mb": self.max_file_size_mb}})
                raise ExcelProcessingError(error_msg)
            
            # Check if file is readable
            try:
                with open(file_path, 'rb') as f:
                    first_kb = f.read(1024)  # Try to read first 1KB
                    metrics.add_metadata("first_kb_size", len(first_kb))
            except (IOError, OSError, PermissionError) as e:
                error_msg = f"Cannot read file {file_path}: {e}"
                metrics.add_metadata("validation_error", "file_not_readable")
                self.logger.error(error_msg, extra={"structured": {"operation": "file_validation_failed", "error_type": "file_not_readable", "file_path": str(file_path), "system_error": str(e)}})
                raise ExcelProcessingError(error_msg)
            
            self.logger.info(f"File validation successful: {file_path} ({file_size_mb:.1f}MB)", extra={"structured": {"operation": "file_validation_success", "file_path": str(file_path), "file_size_mb": file_size_mb, "extension": file_path.suffix.lower()}})
    
    def _extract_worksheets(self, file_path: Path) -> List[WorksheetData]:
        """Extract all worksheets from Excel file.
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            List of WorksheetData objects
        """
        with operation_context(
            "worksheet_extraction",
            self.logger,
            file_path=str(file_path)
        ) as metrics:
            
            worksheets = []
            
            try:
                # First, get worksheet names and metadata using openpyxl
                self.logger.info(f"Reading workbook metadata from {file_path}")
                with operation_context("workbook_metadata_read", self.logger):
                    workbook_info = self._get_workbook_info(file_path)
                
                # Process each worksheet
                with pd.ExcelFile(file_path, engine=None) as excel_file:
                    sheet_names = excel_file.sheet_names
                    sheet_count = len(sheet_names)
                    
                    self.logger.info(
                        f"Found {sheet_count} worksheets in {file_path}: {sheet_names}",
                        extra={
                            "structured": {
                                "operation": "worksheet_discovery",
                                "file_path": str(file_path),
                                "worksheet_count": sheet_count,
                                "worksheet_names": sheet_names
                            }
                        }
                    )
                    
                    metrics.add_metadata("total_worksheets", sheet_count)
                    metrics.add_metadata("worksheet_names", sheet_names)
                    
                    successful_extractions = 0
                    failed_extractions = 0
                    
                    for i, sheet_name in enumerate(sheet_names, 1):
                        with operation_context(
                            "individual_worksheet_processing",
                            self.logger,
                            worksheet_name=sheet_name,
                            file_path=str(file_path),
                            worksheet_index=i,
                            total_worksheets=sheet_count
                        ) as ws_metrics:
                            
                            try:
                                self.logger.info(f"Processing worksheet {i}/{sheet_count}: '{sheet_name}'")
                                
                                worksheet_data = self._process_worksheet(
                                    excel_file, 
                                    sheet_name, 
                                    file_path,
                                    workbook_info.get(sheet_name, {})
                                )
                                
                                if worksheet_data:
                                    worksheets.append(worksheet_data)
                                    successful_extractions += 1
                                    ws_metrics.add_metadata("result", "success")
                                    ws_metrics.add_metadata("row_count", worksheet_data.row_count)
                                    ws_metrics.add_metadata("column_count", worksheet_data.column_count)
                                    
                                    self.logger.info(
                                        f"Successfully processed worksheet '{sheet_name}': "
                                        f"{worksheet_data.row_count:,} rows, {worksheet_data.column_count} columns",
                                        extra={
                                            "structured": {
                                                "operation": "worksheet_processed",
                                                "worksheet_name": sheet_name,
                                                "row_count": worksheet_data.row_count,
                                                "column_count": worksheet_data.column_count,
                                                "has_headers": worksheet_data.has_headers,
                                                "data_types": worksheet_data.column_types
                                            }
                                        }
                                    )
                                else:
                                    failed_extractions += 1
                                    ws_metrics.add_metadata("result", "empty")
                                    self.logger.warning(f"Worksheet '{sheet_name}' was empty or invalid")
                                    
                            except Exception as e:
                                failed_extractions += 1
                                error_type = type(e).__name__
                                ws_metrics.add_metadata("result", "error")
                                ws_metrics.add_metadata("error_type", error_type)
                                
                                self.logger.error(
                                    f"Failed to process worksheet '{sheet_name}': {e}",
                                    extra={
                                        "structured": {
                                            "operation": "worksheet_processing_failed",
                                            "worksheet_name": sheet_name,
                                            "file_path": str(file_path),
                                            "error_type": error_type,
                                            "error_message": str(e)
                                        }
                                    }
                                )
                                # Continue with other worksheets
                                continue
                
                metrics.add_metadata("successful_extractions", successful_extractions)
                metrics.add_metadata("failed_extractions", failed_extractions)
                metrics.add_metadata("success_rate", successful_extractions / sheet_count if sheet_count > 0 else 0)
                
                self.logger.info(
                    f"Worksheet extraction completed: {successful_extractions}/{sheet_count} worksheets processed successfully",
                    extra={
                        "structured": {
                            "operation": "worksheet_extraction_summary",
                            "file_path": str(file_path),
                            "total_worksheets": sheet_count,
                            "successful_extractions": successful_extractions,
                            "failed_extractions": failed_extractions,
                            "success_rate": successful_extractions / sheet_count if sheet_count > 0 else 0
                        }
                    }
                )
                
                return worksheets
                
            except Exception as e:
                raise ExcelProcessingError(f"Failed to extract worksheets: {e}") from e
    
    @log_operation("get_workbook_metadata", log_args=False)
    def _get_workbook_info(self, file_path: Path) -> Dict[str, Dict[str, Any]]:
        """Get workbook metadata using openpyxl.
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Dictionary mapping worksheet names to metadata
        """
        with operation_context(
            "openpyxl_metadata_extraction",
            self.logger,
            file_path=str(file_path)
        ) as metrics:
            
            workbook_info = {}
            
            try:
                # Only process .xlsx files with openpyxl (it doesn't support .xls)
                if file_path.suffix.lower() == '.xlsx':
                    self.logger.debug(f"Extracting workbook metadata using openpyxl for {file_path}")
                    
                    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                    sheet_count = len(workbook.sheetnames)
                    metrics.add_metadata("sheet_count", sheet_count)
                    metrics.add_metadata("extraction_method", "openpyxl")
                    
                    for i, sheet_name in enumerate(workbook.sheetnames, 1):
                        worksheet = workbook[sheet_name]
                        
                        # Get basic worksheet info
                        sheet_info = {
                            'max_row': worksheet.max_row,
                            'max_column': worksheet.max_column,
                            'sheet_state': getattr(worksheet, 'sheet_state', 'visible'),
                            'has_formulas': False,  # Will be determined during pandas processing
                        }
                        workbook_info[sheet_name] = sheet_info
                        
                        self.logger.debug(
                            f"Sheet {i}/{sheet_count} metadata: '{sheet_name}' - {sheet_info['max_row']} rows, {sheet_info['max_column']} columns",
                            extra={
                                "structured": {
                                    "operation": "sheet_metadata_extracted",
                                    "sheet_name": sheet_name,
                                    "sheet_index": i,
                                    "total_sheets": sheet_count,
                                    "max_row": sheet_info['max_row'],
                                    "max_column": sheet_info['max_column'],
                                    "sheet_state": sheet_info['sheet_state']
                                }
                            }
                        )
                    
                    workbook.close()
                    
                    self.logger.info(
                        f"Successfully extracted metadata for {sheet_count} worksheets",
                        extra={
                            "structured": {
                                "operation": "workbook_metadata_success",
                                "file_path": str(file_path),
                                "sheet_count": sheet_count,
                                "extraction_method": "openpyxl"
                            }
                        }
                    )
                    
                else:
                    self.logger.info(f"Skipping openpyxl metadata extraction for {file_path.suffix} file")
                    metrics.add_metadata("extraction_method", "skipped_non_xlsx")
                
            except (InvalidFileException, Exception) as e:
                # Log warning but don't fail - we can still process with pandas
                error_type = type(e).__name__
                metrics.add_metadata("extraction_error", error_type)
                
                self.logger.warning(
                    f"Could not extract metadata with openpyxl: {e}",
                    extra={
                        "structured": {
                            "operation": "workbook_metadata_warning",
                            "file_path": str(file_path),
                            "error_type": error_type,
                            "error_message": str(e),
                            "fallback_available": True
                        }
                    }
                )
            
            return workbook_info
    
    @log_operation("process_individual_worksheet", log_args=False)
    def _process_worksheet(
        self, 
        excel_file: pd.ExcelFile, 
        sheet_name: str,
        source_file: Path,
        metadata: Dict[str, Any]
    ) -> Optional[WorksheetData]:
        """Process individual worksheet and return WorksheetData.
        
        Args:
            excel_file: pandas ExcelFile object
            sheet_name: Name of worksheet to process
            source_file: Source Excel file path
            metadata: Worksheet metadata from openpyxl
            
        Returns:
            WorksheetData object or None if worksheet is empty/invalid
        """
        with operation_context(
            "worksheet_data_processing",
            self.logger,
            worksheet_name=sheet_name,
            source_file=str(source_file)
        ) as metrics:
            
            try:
                # Read worksheet data
                self.logger.debug(f"Reading data for worksheet '{sheet_name}'")
                df = self._read_worksheet_data(excel_file, sheet_name)
                
                # Get basic DataFrame info
                raw_shape = df.shape
                raw_size = df.size
                metrics.add_metadata("raw_shape", raw_shape)
                metrics.add_metadata("raw_size", raw_size)
                
                # Skip completely empty worksheets
                if df.empty:
                    self.logger.info(
                        f"Skipping empty worksheet: {sheet_name}",
                        extra={
                            "structured": {
                                "operation": "worksheet_skipped",
                                "worksheet_name": sheet_name,
                                "reason": "completely_empty",
                                "source_file": str(source_file)
                            }
                        }
                    )
                    metrics.add_metadata("skip_reason", "completely_empty")
                    return None
                
                # Skip worksheets with no actual data
                if df.dropna(how='all').empty:
                    self.logger.info(
                        f"Skipping worksheet with no data: {sheet_name}",
                        extra={
                            "structured": {
                                "operation": "worksheet_skipped",
                                "worksheet_name": sheet_name,
                                "reason": "no_actual_data",
                                "source_file": str(source_file),
                                "raw_shape": raw_shape
                            }
                        }
                    )
                    metrics.add_metadata("skip_reason", "no_actual_data")
                    return None
                
                # Enhance metadata with pandas-derived info
                self.logger.debug(f"Enhancing metadata for worksheet '{sheet_name}'")
                enhanced_metadata = self._enhance_metadata(df, metadata)
                
                # Create WorksheetData object
                worksheet_data = WorksheetData(
                    source_file=source_file,
                    worksheet_name=sheet_name,
                    data=df,
                    metadata=enhanced_metadata
                )
                
                # Log detailed processing results
                self.logger.info(
                    f"Successfully processed worksheet '{sheet_name}': "
                    f"{worksheet_data.row_count:,} rows, "
                    f"{worksheet_data.column_count} columns, "
                    f"density: {worksheet_data.data_density:.3f}",
                    extra={
                        "structured": {
                            "operation": "worksheet_processing_complete",
                            "worksheet_name": sheet_name,
                            "source_file": str(source_file),
                            "row_count": worksheet_data.row_count,
                            "column_count": worksheet_data.column_count,
                            "data_density": worksheet_data.data_density,
                            "has_headers": worksheet_data.has_headers,
                            "column_types": worksheet_data.column_types,
                            "metadata_keys": list(enhanced_metadata.keys())
                        }
                    }
                )
                
                # Add comprehensive metrics
                metrics.add_metadata("result", "success")
                metrics.add_metadata("row_count", worksheet_data.row_count)
                metrics.add_metadata("column_count", worksheet_data.column_count)
                metrics.add_metadata("data_density", worksheet_data.data_density)
                metrics.add_metadata("has_headers", worksheet_data.has_headers)
                metrics.add_metadata("total_cells", worksheet_data.row_count * worksheet_data.column_count)
                
                return worksheet_data
                
            except Exception as e:
                error_type = type(e).__name__
                error_msg = f"Failed to read worksheet '{sheet_name}': {e}"
                
                metrics.add_metadata("result", "error")
                metrics.add_metadata("error_type", error_type)
                
                self.logger.error(
                    error_msg,
                    exc_info=True,
                    extra={
                        "structured": {
                            "operation": "worksheet_processing_failed",
                            "worksheet_name": sheet_name,
                            "source_file": str(source_file),
                            "error_type": error_type,
                            "error_message": str(e)
                        }
                    }
                )
                return None
    
    @log_operation("read_worksheet_data", log_args=False)
    def _read_worksheet_data(self, excel_file: pd.ExcelFile, sheet_name: str) -> pd.DataFrame:
        """Read worksheet data using pandas.
        
        Args:
            excel_file: pandas ExcelFile object
            sheet_name: Name of worksheet to read
            
        Returns:
            DataFrame containing worksheet data
        """
        with operation_context(
            "pandas_worksheet_read",
            self.logger,
            worksheet_name=sheet_name
        ) as metrics:
            
            try:
                self.logger.debug(f"Reading worksheet data with pandas: '{sheet_name}'")
                
                # Read the worksheet
                df = pd.read_excel(
                    excel_file,
                    sheet_name=sheet_name,
                    header=None,  # Don't assume first row is header
                    na_values=[''],  # Treat empty strings as NaN
                    keep_default_na=True,
                    na_filter=True,
                    dtype=str,  # Read everything as string initially
                    engine=None  # Let pandas choose the best engine
                )
                
                # Log read results
                shape = df.shape
                size = df.size
                non_null_count = df.count().sum()
                
                self.logger.debug(
                    f"Worksheet data read successfully: '{sheet_name}' - {shape[0]} rows x {shape[1]} columns, {non_null_count:,} non-null cells",
                    extra={
                        "structured": {
                            "operation": "worksheet_data_read",
                            "worksheet_name": sheet_name,
                            "shape": shape,
                            "total_cells": size,
                            "non_null_cells": non_null_count,
                            "data_ratio": non_null_count / size if size > 0 else 0
                        }
                    }
                )
                
                metrics.add_metadata("shape", shape)
                metrics.add_metadata("total_cells", size)
                metrics.add_metadata("non_null_cells", non_null_count)
                metrics.add_metadata("read_success", True)
                
                return df
                
            except Exception as e:
                error_type = type(e).__name__
                metrics.add_metadata("read_success", False)
                metrics.add_metadata("error_type", error_type)
                
                self.logger.error(
                    f"Failed to read worksheet data for '{sheet_name}': {e}",
                    extra={
                        "structured": {
                            "operation": "worksheet_data_read_failed",
                            "worksheet_name": sheet_name,
                            "error_type": error_type,
                            "error_message": str(e)
                        }
                    }
                )
                raise ExcelProcessingError(f"Failed to read worksheet data: {e}") from e
    
    @log_operation("enhance_worksheet_metadata", log_args=False)
    def _enhance_metadata(self, df: pd.DataFrame, base_metadata: Dict[str, Any]) -> Dict[str, Any]:
        """Enhance metadata with pandas-derived information.
        
        Args:
            df: DataFrame containing worksheet data
            base_metadata: Base metadata from openpyxl
            
        Returns:
            Enhanced metadata dictionary
        """
        with operation_context(
            "metadata_enhancement",
            self.logger,
            base_metadata_keys=list(base_metadata.keys()),
            dataframe_shape=df.shape
        ) as metrics:
            
            metadata = base_metadata.copy()
            
            # Basic data info
            metadata.update({
                'pandas_shape': df.shape,
                'pandas_columns': list(df.columns),
                'pandas_dtypes': {str(col): str(dtype) for col, dtype in df.dtypes.items()},
                'total_cells': df.size,
                'non_empty_cells': df.count().sum(),
                'empty_cells': df.isnull().sum().sum(),
                'data_density': (df.count().sum() / df.size) if df.size > 0 else 0.0,
            })
        
            # Data type analysis
            numeric_columns = []
            date_columns = []
            text_columns = []
            
            for col in df.columns:
                # Sample non-null values for type detection
                sample_data = df[col].dropna().head(100)
                
                if len(sample_data) == 0:
                    continue
                
                # Try to detect numeric columns
                try:
                    pd.to_numeric(sample_data, errors='raise')
                    numeric_columns.append(col)
                    continue
                except (ValueError, TypeError):
                    pass
                
                # Try to detect date columns
                try:
                    import warnings
                    with warnings.catch_warnings():
                        warnings.simplefilter("ignore", UserWarning)
                        pd.to_datetime(sample_data, errors='raise')
                    date_columns.append(col)
                    continue
                except (ValueError, TypeError):
                    pass
                
                # Default to text
                text_columns.append(col)
        
            metadata.update({
                'numeric_columns': numeric_columns,
                'date_columns': date_columns, 
                'text_columns': text_columns,
                'column_types_detected': {
                    'numeric': len(numeric_columns),
                    'date': len(date_columns),
                    'text': len(text_columns),
                }
            })
            
            # Row analysis
            empty_rows = df.isnull().all(axis=1).sum()
            full_rows = df.notnull().all(axis=1).sum()
            
            metadata.update({
                'empty_rows': empty_rows,
                'full_rows': full_rows,
                'partial_rows': len(df) - empty_rows - full_rows,
            })
            
            # Column analysis
            empty_columns = df.isnull().all(axis=0).sum()
            full_columns = df.notnull().all(axis=0).sum()
            
            metadata.update({
                'empty_columns': empty_columns,
                'full_columns': full_columns,
                'partial_columns': len(df.columns) - empty_columns - full_columns,
            })
            
            # Potential header detection (simple heuristic)
            if len(df) > 0:
                first_row_text_ratio = 0
                if len(df.columns) > 0:
                    first_row = df.iloc[0]
                    text_cells = sum(1 for val in first_row if isinstance(val, str) and val.strip())
                    first_row_text_ratio = text_cells / len(first_row)
                
                metadata['potential_header_row'] = 0 if first_row_text_ratio > 0.5 else None
                metadata['first_row_text_ratio'] = first_row_text_ratio
            
            # Log metadata enhancement completion
            enhancement_count = len(metadata) - len(base_metadata)
            metrics.add_metadata("enhancement_count", enhancement_count)
            metrics.add_metadata("final_metadata_keys", list(metadata.keys()))
            
            self.logger.debug(
                f"Metadata enhancement completed: {len(base_metadata)} -> {len(metadata)} fields (+{enhancement_count})",
                extra={
                    "structured": {
                        "operation": "metadata_enhancement_complete",
                        "original_fields": len(base_metadata),
                        "enhanced_fields": len(metadata),
                        "new_fields": enhancement_count,
                        "data_density": metadata.get('data_density', 0),
                        "column_types_detected": metadata.get('column_types_detected', {})
                    }
                }
            )
            
            return metadata
    
    @log_operation("get_worksheet_preview", log_args=False)
    def get_worksheet_preview(
        self, 
        file_path: Union[str, Path], 
        sheet_name: str,
        max_rows: int = 10
    ) -> pd.DataFrame:
        """Get preview of worksheet data.
        
        Args:
            file_path: Path to Excel file
            sheet_name: Name of worksheet
            max_rows: Maximum number of rows to return
            
        Returns:
            DataFrame with preview data
            
        Raises:
            ExcelProcessingError: If preview cannot be generated
        """
        file_path = Path(file_path)
        
        with operation_context(
            "worksheet_preview_generation",
            self.logger,
            file_path=str(file_path),
            worksheet_name=sheet_name,
            max_rows=max_rows
        ) as metrics:
            
            try:
                self.logger.info(f"Generating preview for worksheet '{sheet_name}' (max {max_rows} rows)")
                
                with pd.ExcelFile(file_path) as excel_file:
                    df = pd.read_excel(
                        excel_file,
                        sheet_name=sheet_name,
                        header=None,
                        nrows=max_rows,
                        dtype=str
                    )
                    
                    # Log preview results
                    actual_rows = len(df)
                    actual_cols = len(df.columns) if not df.empty else 0
                    
                    metrics.add_metadata("preview_rows", actual_rows)
                    metrics.add_metadata("preview_columns", actual_cols)
                    metrics.add_metadata("requested_max_rows", max_rows)
                    
                    self.logger.info(
                        f"Preview generated successfully: {actual_rows} rows x {actual_cols} columns",
                        extra={
                            "structured": {
                                "operation": "worksheet_preview_success",
                                "file_path": str(file_path),
                                "worksheet_name": sheet_name,
                                "preview_rows": actual_rows,
                                "preview_columns": actual_cols,
                                "max_rows_requested": max_rows
                            }
                        }
                    )
                    
                    return df
                    
            except Exception as e:
                error_type = type(e).__name__
                metrics.add_metadata("error_type", error_type)
                
                self.logger.error(
                    f"Failed to generate preview for '{sheet_name}': {e}",
                    extra={
                        "structured": {
                            "operation": "worksheet_preview_failed",
                            "file_path": str(file_path),
                            "worksheet_name": sheet_name,
                            "max_rows": max_rows,
                            "error_type": error_type,
                            "error_message": str(e)
                        }
                    }
                )
                raise ExcelProcessingError(f"Failed to generate preview: {e}") from e
    
    @log_operation("list_worksheets", log_args=False)
    def list_worksheets(self, file_path: Union[str, Path]) -> List[str]:
        """List all worksheet names in Excel file.
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            List of worksheet names
            
        Raises:
            ExcelProcessingError: If worksheets cannot be listed
        """
        file_path = Path(file_path)
        
        with operation_context(
            "worksheet_listing",
            self.logger,
            file_path=str(file_path)
        ) as metrics:
            
            try:
                self.logger.debug(f"Listing worksheets in {file_path}")
                
                with pd.ExcelFile(file_path) as excel_file:
                    sheet_names = excel_file.sheet_names
                    sheet_count = len(sheet_names)
                    
                    metrics.add_metadata("worksheet_count", sheet_count)
                    metrics.add_metadata("worksheet_names", sheet_names)
                    
                    self.logger.info(
                        f"Found {sheet_count} worksheets: {sheet_names}",
                        extra={
                            "structured": {
                                "operation": "worksheet_listing_success",
                                "file_path": str(file_path),
                                "worksheet_count": sheet_count,
                                "worksheet_names": sheet_names
                            }
                        }
                    )
                    
                    return sheet_names
                    
            except Exception as e:
                error_type = type(e).__name__
                metrics.add_metadata("error_type", error_type)
                
                self.logger.error(
                    f"Failed to list worksheets in {file_path}: {e}",
                    extra={
                        "structured": {
                            "operation": "worksheet_listing_failed",
                            "file_path": str(file_path),
                            "error_type": error_type,
                            "error_message": str(e)
                        }
                    }
                )
                raise ExcelProcessingError(f"Failed to list worksheets: {e}") from e