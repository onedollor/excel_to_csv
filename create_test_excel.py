#!/usr/bin/env python3
"""
Create comprehensive test Excel files for CSV conversion testing.

This script creates test files covering various scenarios:
- Problematic worksheets that cannot/should not be converted
- Normal worksheets that can be converted successfully  
- Edge cases and boundary conditions
- Different data types and formats
"""

import pandas as pd
import numpy as np
from pathlib import Path
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
warnings.filterwarnings('ignore')


def create_problematic_worksheet(wb):
    """Create a worksheet that should NOT be converted to CSV."""
    ws = wb.create_sheet("Problematic_Sheet")
    
    # Add a header
    ws['A1'] = "❌ PROBLEMATIC WORKSHEET - NOT SUITABLE FOR CSV"
    ws['A1'].font = Font(bold=True, color="FF0000")
    ws['A1'].fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    
    # Scenario 1: Completely empty worksheet (after header)
    # This will make it appear to have content but actually be empty of data
    
    # Scenario 2: Merged cells that would break CSV structure
    ws.merge_cells('A3:D3')
    ws['A3'] = "This is a merged cell spanning multiple columns"
    ws['A3'].alignment = Alignment(horizontal='center')
    
    # Scenario 3: Complex formatting that doesn't translate to CSV
    ws.merge_cells('A5:B6') 
    ws['A5'] = "Complex\nMulti-line\nCell"
    ws['A5'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Scenario 4: Formulas referencing external worksheets
    ws['A8'] = "=SUM(Normal_Sheet!A1:A10)"
    ws['B8'] = "=IF(ISBLANK(A8), 'No Data', A8)"
    ws['C8'] = "=TODAY()"
    ws['D8'] = "=RAND()"
    
    # Scenario 5: Very sparse data that might indicate non-tabular content
    ws['A15'] = "Scattered"
    ws['G20'] = "Data"
    ws['C25'] = "Points"
    ws['J30'] = "Here"
    
    # Scenario 6: Images and charts metadata (simulated with text)
    ws['A35'] = "[IMAGE: chart1.png]"
    ws['A36'] = "[CHART: Sales Data Q1-Q4]"
    ws['A37'] = "[DRAWING: Process Flow Diagram]"
    
    # Scenario 7: Headers without consistent data structure
    headers = ["Name", "", "Value", "", "", "Notes", "Status"]
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=40, column=i, value=header)
        cell.font = Font(bold=True)
    
    # Irregular data under headers
    ws['A41'] = "Item1"
    ws['C41'] = 100
    ws['G41'] = "OK"
    ws['A43'] = "Item2"  # Note: row 42 is skipped
    ws['F43'] = "Some notes here"
    
    return ws


def create_normal_worksheet(wb):
    """Create a normal worksheet that CAN be converted to CSV successfully."""
    ws = wb.create_sheet("Normal_Sheet")
    
    # Create a proper tabular dataset
    data = {
        'ID': range(1, 101),
        'Name': [f'Product_{i:03d}' for i in range(1, 101)],
        'Category': np.random.choice(['Electronics', 'Clothing', 'Books', 'Home', 'Sports'], 100),
        'Price': np.round(np.random.uniform(10, 500, 100), 2),
        'Quantity': np.random.randint(1, 100, 100),
        'Date_Added': [datetime.date(2024, 1, 1) + datetime.timedelta(days=int(x)) for x in np.random.randint(0, 365, 100)],
        'In_Stock': np.random.choice([True, False], 100),
        'Rating': np.round(np.random.uniform(1, 5, 100), 1),
        'Description': [f'High quality product item {i} with excellent features' for i in range(1, 101)]
    }
    
    df = pd.DataFrame(data)
    
    # Add header row with formatting
    headers = list(df.columns)
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    
    # Add data rows
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    return ws


def create_edge_case_worksheet(wb):
    """Create worksheet with edge cases that CAN still be converted."""
    ws = wb.create_sheet("Edge_Cases_Sheet")
    
    # Headers
    headers = [
        'Text_Data', 'Numbers', 'Dates', 'Special_Chars', 'Empty_Cells',
        'Long_Text', 'Unicode', 'Formulas', 'Boolean', 'Mixed_Types'
    ]
    
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=header)
        cell.font = Font(bold=True)
    
    # Edge case data
    edge_cases = [
        # Row 2: Normal data
        ['Normal Text', 42, datetime.date(2024, 1, 15), 'Standard', 'Value', 'Short text', 'ASCII', '=2+2', True, 'Mixed'],
        
        # Row 3: Special characters and encoding
        ['Text with "quotes"', -123.45, datetime.date(2024, 12, 31), 'Special: !@#$%^&*()', '', 
         'Very long text ' * 20, 'Unicode: 中文 русский العربية 🎉', '=A3&B3', False, 123.45],
        
        # Row 4: Empty and None values
        ['', None, None, '', None, '', '', '', None, ''],
        
        # Row 5: Extreme values
        ['MIN_VALUE', -999999999, datetime.date(1900, 1, 1), '|\\/<>', 'NULL', 
         'X' * 500, '♠♣♥♦→↓←↑', '=NOW()', True, float('inf')],
        
        # Row 6: Boundary cases
        ['Newline\nText', 0, datetime.date.today(), 'Comma,Semicolon;', 'N/A',
         'Tab\tSeparated\tText', 'Émojis: 🚀🌟⭐', '=RAND()', False, -0.0],
        
        # Row 7: CSV problematic characters
        ['Text, with, commas', 3.14159, datetime.date(2025, 1, 1), '"Quoted"Values"', 'BLANK',
         'Line\nBreak\rReturn', 'åäö ñüÿ', '=PI()', True, 'String123'],
        
        # Row 8: Scientific notation and precision
        ['Scientific', 1.23e-10, datetime.date(2024, 6, 15), '~!@#$%^&*()_+', '',
         'Precision test', '∑∏∆∇', '=EXP(1)', False, 1.23456789012345],
        
        # Row 9: Boolean and null-like strings
        ['TRUE', 0.0001, datetime.date(2024, 3, 29), 'NULL', 'None',
         'Empty string test', 'null', '=FALSE()', None, 'true'],
        
        # Row 10: Mixed numeric formats
        ['Currency', 1000000, datetime.date(2024, 7, 4), '$1,234.56', 'ZERO',
         'Percentage test', '100%', '=1/3', True, '1.23E+10']
    ]
    
    for r_idx, row_data in enumerate(edge_cases, 2):
        for c_idx, value in enumerate(row_data, 1):
            try:
                ws.cell(row=r_idx, column=c_idx, value=value)
            except Exception as e:
                # Handle any values that can't be written to Excel
                ws.cell(row=r_idx, column=c_idx, value=str(value))
    
    return ws


def create_additional_test_worksheets(wb):
    """Create additional worksheets for comprehensive testing."""
    
    # Worksheet 4: Very large dataset
    ws_large = wb.create_sheet("Large_Dataset")
    ws_large['A1'] = "Large Dataset Test (1000+ rows)"
    ws_large['A1'].font = Font(bold=True)
    
    headers = ['Col_A', 'Col_B', 'Col_C', 'Col_D', 'Col_E']
    for i, header in enumerate(headers, 1):
        cell = ws_large.cell(row=2, column=i, value=header)
        cell.font = Font(bold=True)
    
    # Generate large dataset
    for row in range(3, 1003):  # 1000 rows of data
        for col in range(1, 6):
            ws_large.cell(row=row, column=col, value=f'Data_{row}_{col}')
    
    # Worksheet 5: Only headers (no data)
    ws_headers_only = wb.create_sheet("Headers_Only")
    headers = ['Name', 'Value', 'Status', 'Date']
    for i, header in enumerate(headers, 1):
        cell = ws_headers_only.cell(row=1, column=i, value=header)
        cell.font = Font(bold=True)
    # No data rows - just headers
    
    # Worksheet 6: Single column data
    ws_single = wb.create_sheet("Single_Column")
    ws_single['A1'] = "Single_Column_Data"
    ws_single['A1'].font = Font(bold=True)
    
    for i in range(2, 52):  # 50 rows
        ws_single.cell(row=i, column=1, value=f'Value_{i-1}')
    
    # Worksheet 7: Wide dataset (many columns)
    ws_wide = wb.create_sheet("Wide_Dataset")
    
    # Create 50 columns
    wide_headers = [f'Col_{chr(65 + i//26)}{chr(65 + i%26)}' for i in range(50)]
    for i, header in enumerate(wide_headers, 1):
        cell = ws_wide.cell(row=1, column=i, value=header)
        cell.font = Font(bold=True)
    
    # Add 10 rows of data
    for row in range(2, 12):
        for col in range(1, 51):
            ws_wide.cell(row=row, column=col, value=f'R{row}C{col}')
    
    # Worksheet 8: Mixed data types per column
    ws_mixed = wb.create_sheet("Mixed_Types")
    
    headers = ['Mixed_Col_1', 'Mixed_Col_2', 'Mixed_Col_3']
    for i, header in enumerate(headers, 1):
        cell = ws_mixed.cell(row=1, column=i, value=header)
        cell.font = Font(bold=True)
    
    mixed_data = [
        [1, 'Text', True],
        ['String', 2.5, False],
        [datetime.date.today(), 100, 'Yes'],
        [3.14, datetime.date(2024, 1, 1), None],
        ['Mixed', None, 42]
    ]
    
    for r_idx, row_data in enumerate(mixed_data, 2):
        for c_idx, value in enumerate(row_data, 1):
            ws_mixed.cell(row=r_idx, column=c_idx, value=value)
    
    return wb


def create_comprehensive_test_file():
    """Create the main comprehensive test Excel file."""
    
    print("Creating comprehensive test Excel file...")
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create the three main test worksheets
    print("  Creating problematic worksheet...")
    create_problematic_worksheet(wb)
    
    print("  Creating normal worksheet...")
    create_normal_worksheet(wb)
    
    print("  Creating edge cases worksheet...")
    create_edge_case_worksheet(wb)
    
    print("  Creating additional test worksheets...")
    create_additional_test_worksheets(wb)
    
    # Add a summary/info sheet
    ws_info = wb.create_sheet("Test_Info", 0)  # Insert at beginning
    
    ws_info['A1'] = "EXCEL TO CSV CONVERSION TEST FILE"
    ws_info['A1'].font = Font(bold=True, size=16)
    
    info_content = [
        "",
        "This file contains multiple worksheets for testing CSV conversion:",
        "",
        "📋 WORKSHEET DESCRIPTIONS:",
        "",
        "1. Problematic_Sheet - ❌ Should NOT convert to CSV",
        "   - Contains merged cells, complex formatting, sparse data",
        "   - Formulas, images, non-tabular structure",
        "",
        "2. Normal_Sheet - ✅ Should convert successfully",
        "   - Well-structured tabular data with 100 rows",
        "   - Multiple data types (text, numbers, dates, booleans)",
        "",
        "3. Edge_Cases_Sheet - ✅ Should convert with special handling",
        "   - Special characters, Unicode, empty cells",
        "   - CSV problematic characters (quotes, commas, newlines)",
        "",
        "4. Large_Dataset - ✅ Performance test (1000+ rows)",
        "",
        "5. Headers_Only - ⚠️  Edge case (headers but no data)",
        "",
        "6. Single_Column - ✅ Minimal valid dataset",
        "",
        "7. Wide_Dataset - ✅ Many columns test (50 columns)",
        "",
        "8. Mixed_Types - ✅ Mixed data types per column",
        "",
        "🧪 TEST SCENARIOS COVERED:",
        "- Normal tabular data conversion",
        "- Problematic sheets that should be skipped",
        "- Edge cases and boundary conditions", 
        "- Performance with large datasets",
        "- Various data types and encodings",
        "- CSV-problematic characters",
        "- Empty/sparse data handling",
        "- Wide and narrow datasets",
        "",
        f"📅 Created: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ]
    
    for i, line in enumerate(info_content, 2):
        ws_info[f'A{i}'] = line
        if line.startswith(('📋', '🧪')):
            ws_info[f'A{i}'].font = Font(bold=True)
    
    # Save the file
    output_path = Path("/tmp/comprehensive_test_file.xlsx")
    wb.save(output_path)
    
    print(f"✅ Comprehensive test file created: {output_path}")
    print(f"   File size: {output_path.stat().st_size:,} bytes")
    print(f"   Worksheets: {len(wb.sheetnames)}")
    print(f"   Worksheet names: {', '.join(wb.sheetnames)}")
    
    return output_path


def create_simple_test_files():
    """Create individual simple test files for specific scenarios."""
    
    test_files = {}
    
    # 1. Completely empty Excel file
    print("Creating empty Excel file...")
    wb_empty = Workbook()
    wb_empty.remove(wb_empty['Sheet'])  # Remove default sheet
    wb_empty.create_sheet("Empty_Sheet")  # Add empty sheet
    empty_path = Path("/tmp/empty_test.xlsx")
    wb_empty.save(empty_path)
    test_files['empty'] = empty_path
    
    # 2. Excel file with only formulas
    print("Creating formulas-only Excel file...")
    wb_formulas = Workbook()
    ws = wb_formulas.active
    ws.title = "Formulas_Only"
    ws['A1'] = "Formula_Results"
    ws['A2'] = "=1+1"
    ws['A3'] = "=TODAY()"
    ws['A4'] = "=RAND()"
    ws['A5'] = "=PI()"
    formulas_path = Path("/tmp/formulas_test.xlsx")
    wb_formulas.save(formulas_path)
    test_files['formulas'] = formulas_path
    
    # 3. Excel file with corrupted/problematic data
    print("Creating problematic data Excel file...")
    wb_problem = Workbook()
    ws = wb_problem.active
    ws.title = "Problematic_Data"
    
    # Add data that might cause issues
    problematic_data = [
        ['Header1', 'Header2', 'Header3'],
        ['Normal', 'Data', 'Row'],
        [None, '', '   '],  # Empty/null row
        ['Very' * 1000, 'Long' * 1000, 'Text' * 1000],  # Extremely long text
        [float('inf'), float('-inf'), float('nan')],  # Infinity and NaN
        ['Multi\nLine\nText', 'Tab\tSeparated', 'Quote"Inside'],  # Special chars
    ]
    
    for r_idx, row in enumerate(problematic_data, 1):
        for c_idx, value in enumerate(row, 1):
            try:
                ws.cell(row=r_idx, column=c_idx, value=value)
            except:
                ws.cell(row=r_idx, column=c_idx, value=str(value)[:1000])  # Truncate if needed
    
    problem_path = Path("/tmp/problematic_test.xlsx")
    wb_problem.save(problem_path)
    test_files['problematic'] = problem_path
    
    return test_files


if __name__ == "__main__":
    print("🧪 Creating comprehensive Excel test files for CSV conversion testing\n")
    
    # Create main comprehensive test file
    main_file = create_comprehensive_test_file()
    
    print("\n" + "="*60)
    
    # Create additional simple test files
    print("Creating additional specific test files...")
    simple_files = create_simple_test_files()
    
    print(f"\n✅ All test files created successfully!")
    print(f"\n📁 Test Files Created:")
    print(f"   Main comprehensive test: {main_file}")
    for name, path in simple_files.items():
        print(f"   {name.capitalize()} test: {path}")
    
    print(f"\n🎯 Ready for testing the Excel to CSV conversion system!")
    print("   Use these files to test various conversion scenarios and edge cases.")