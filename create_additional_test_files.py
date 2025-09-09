#!/usr/bin/env python3
"""
Create specific additional test Excel files for comprehensive testing.

This script creates three specific test files:
1. empty_test.xlsx - Excel file with no worksheets  
2. formulas_test.xlsx - Excel file with formula-only content
3. problematic_test.xlsx - Excel file with problematic data
"""

import pandas as pd
import numpy as np
from pathlib import Path
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
warnings.filterwarnings('ignore')


def create_empty_test_file():
    """Create an Excel file with no worksheets (or empty worksheets)."""
    print("Creating empty_test.xlsx...")
    
    # Create workbook and remove all default sheets
    wb = Workbook()
    
    # Remove the default 'Sheet' worksheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Option 1: Completely empty workbook (no worksheets)
    # This might cause issues, so let's create an empty worksheet instead
    
    # Option 2: Create an empty worksheet with no data
    ws = wb.create_sheet("Empty_Sheet")
    # Don't add any data - completely empty
    
    output_path = Path("/tmp/empty_test.xlsx")
    wb.save(output_path)
    
    print(f"✅ Created empty test file: {output_path}")
    print(f"   File size: {output_path.stat().st_size:,} bytes")
    print(f"   Worksheets: {len(wb.sheetnames)}")
    
    return output_path


def create_formulas_test_file():
    """Create an Excel file with formula-only content."""
    print("Creating formulas_test.xlsx...")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Formulas_Only"
    
    # Add header
    ws['A1'] = "Formula-Only Worksheet"
    ws['A1'].font = Font(bold=True, color="FF0000")
    
    # Row 2: Headers for formula results
    headers = ['Formula_Type', 'Formula', 'Result_Type', 'Notes']
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=i, value=header)
        cell.font = Font(bold=True)
    
    # Add various types of formulas
    formula_data = [
        # Row 3: Basic arithmetic
        ['Arithmetic', '=2+3*4', 'Number', 'Basic math operations'],
        
        # Row 4: Date functions
        ['Date', '=TODAY()', 'Date', 'Current date function'],
        
        # Row 5: Text functions  
        ['Text', '=UPPER("hello world")', 'Text', 'Text manipulation'],
        
        # Row 6: Statistical functions
        ['Statistics', '=AVERAGE(1,2,3,4,5)', 'Number', 'Statistical calculation'],
        
        # Row 7: Random functions
        ['Random', '=RAND()', 'Number', 'Random number generation'],
        
        # Row 8: Logical functions
        ['Logic', '=IF(5>3,"TRUE","FALSE")', 'Text', 'Conditional logic'],
        
        # Row 9: Mathematical functions
        ['Math', '=PI()*2', 'Number', 'Mathematical constants'],
        
        # Row 10: Nested formulas
        ['Nested', '=ROUND(SQRT(16),2)', 'Number', 'Nested function calls'],
        
        # Row 11: String concatenation
        ['Concat', '=CONCATENATE("Hello"," ","World")', 'Text', 'String joining'],
        
        # Row 12: Complex formula
        ['Complex', '=SUM(1:10)+AVERAGE(1:5)*COUNT(1:3)', 'Number', 'Multi-function formula'],
    ]
    
    for row_idx, (formula_type, formula, result_type, notes) in enumerate(formula_data, 3):
        ws.cell(row=row_idx, column=1, value=formula_type)
        ws.cell(row=row_idx, column=2, value=formula)  # This will be stored as formula
        ws.cell(row=row_idx, column=3, value=result_type)
        ws.cell(row=row_idx, column=4, value=notes)
    
    # Add some cells that are pure formulas (no descriptive text)
    ws['F1'] = '=NOW()'
    ws['F2'] = '=RAND()'
    ws['F3'] = '=1+2+3+4+5'
    ws['F4'] = '=LEN("Test String")'
    ws['F5'] = '=POWER(2,8)'
    
    # Add cross-referencing formulas
    ws['G1'] = '=F1+1'  # Reference to another formula
    ws['G2'] = '=SUM(F1:F5)'  # Sum of other formulas
    
    output_path = Path("/tmp/formulas_test.xlsx")
    wb.save(output_path)
    
    print(f"✅ Created formulas test file: {output_path}")
    print(f"   File size: {output_path.stat().st_size:,} bytes")
    print(f"   Worksheets: {len(wb.sheetnames)}")
    print(f"   Formula cells: ~20 cells with formulas")
    
    return output_path


def create_problematic_test_file():
    """Create an Excel file with problematic data that's hard to convert to CSV."""
    print("Creating problematic_test.xlsx...")
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Problematic_Data"
    
    # Add header
    ws['A1'] = "Problematic Data Test"
    ws['A1'].font = Font(bold=True, color="FF0000")
    
    # Headers row
    headers = ['Mixed_Data', 'Extreme_Values', 'Special_Characters']
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=i, value=header)
        cell.font = Font(bold=True)
    
    # Problematic data that should cause issues or low confidence scores
    problematic_data = [
        # Row 3: Mixed types in each column
        ['String Value', 42, 'Normal Text'],
        
        # Row 4: Extreme values and special cases
        [12345, float('inf'), 'Text with "quotes" and commas, semicolons;'],
        
        # Row 5: Empty and None-like values
        [None, '', 'NULL'],
        
        # Row 6: Very long text that might break CSV
        ['Very long text ' * 100, -999999999, 'Multi\nLine\nText\nWith\nBreaks'],
        
        # Row 7: Special numeric representations
        ['NaN', float('nan'), '1.23E+45'],
        
        # Row 8: Unicode and special characters
        ['Unicode: 中文 русский العربية 🎉', -float('inf'), 'Symbols: ♠♣♥♦→↓←↑'],
        
        # Row 9: CSV problematic content
        ['Contains,commas,everywhere', 'Contains\ttabs\there', 'Contains"quotes"inside'],
        
        # Row 10: Binary-like content (safe representations)
        ['Binary content (bytes)', '\\x00\\x01\\x02 (escaped)', 'File\\x00Path\\x00Here (escaped)'],
        
        # Row 11: Boolean variations
        [True, False, 'true'],
        
        # Row 12: Date-like but inconsistent
        ['2024-01-15', 'Jan 15, 2024', '15/01/24'],
    ]
    
    for row_idx, row_data in enumerate(problematic_data, 3):
        for col_idx, value in enumerate(row_data, 1):
            try:
                ws.cell(row=row_idx, column=col_idx, value=value)
            except Exception as e:
                # If value can't be written, convert to string
                ws.cell(row=row_idx, column=col_idx, value=str(value))
    
    # Add some scattered data in random positions (sparse data pattern)
    ws['F1'] = 'Scattered'
    ws['H5'] = 'Random'
    ws['B10'] = 'Sparse'
    ws['J15'] = 'Data'
    ws['D20'] = 'Points'
    
    # Add merged cells (problematic for CSV)
    ws.merge_cells('E1:G1')
    ws['E1'] = "Merged Cell Content"
    ws['E1'].alignment = Alignment(horizontal='center')
    
    ws.merge_cells('E2:F3')
    ws['E2'] = "Another\nMerged\nArea"
    ws['E2'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Add some formulas mixed with data
    ws['K1'] = '=RAND()'
    ws['K2'] = '=TODAY()'
    ws['K3'] = 'Mixed Content'
    ws['K4'] = 42
    
    output_path = Path("/tmp/problematic_test.xlsx")
    wb.save(output_path)
    
    print(f"✅ Created problematic test file: {output_path}")
    print(f"   File size: {output_path.stat().st_size:,} bytes")
    print(f"   Worksheets: {len(wb.sheetnames)}")
    print(f"   Contains: Mixed data types, extreme values, merged cells, sparse data")
    
    return output_path


def create_all_additional_test_files():
    """Create all three additional test files."""
    print("🧪 Creating Additional Test Files for Excel to CSV Testing\n")
    
    test_files = {}
    
    # Create each test file
    test_files['empty'] = create_empty_test_file()
    print()
    
    test_files['formulas'] = create_formulas_test_file()
    print()
    
    test_files['problematic'] = create_problematic_test_file()
    print()
    
    print("=" * 60)
    print("✅ All additional test files created successfully!")
    print("\n📁 Test Files Summary:")
    
    for name, path in test_files.items():
        if path.exists():
            size = path.stat().st_size
            print(f"   {name.capitalize()}: {path.name} ({size:,} bytes)")
        else:
            print(f"   {name.capitalize()}: {path.name} (FILE NOT FOUND)")
    
    print(f"\n🎯 Ready for testing with the enhanced logging system!")
    print("   These files test specific edge cases:")
    print("   - empty_test.xlsx: Tests handling of empty/minimal content")
    print("   - formulas_test.xlsx: Tests formula-heavy worksheets")
    print("   - problematic_test.xlsx: Tests problematic data structures")
    
    return test_files


if __name__ == "__main__":
    create_all_additional_test_files()