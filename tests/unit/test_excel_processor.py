"""Comprehensive unit tests for Excel Processor - Fixed Version.

This test suite provides comprehensive coverage for:
- Excel file reading and processing (actual methods)
- Format validation and error handling
- Multi-sheet processing
- Memory management and large file handling
"""

import pytest
from pathlib import Path
from unittest.mock import Mock, patch, MagicMock
from typing import List, Dict, Any

import pandas as pd
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

from excel_to_csv.processors.excel_processor import ExcelProcessor, ExcelProcessingError
from excel_to_csv.models.data_models import WorksheetData


class TestExcelProcessor:
    """Test suite for ExcelProcessor initialization and basic functionality."""

    @pytest.fixture
    def excel_processor(self):
        """Create a fresh ExcelProcessor instance for testing."""
        return ExcelProcessor(max_file_size_mb=100, chunk_size=10000)

    def test_init_default_parameters(self):
        """Test ExcelProcessor initialization with default parameters."""
        processor = ExcelProcessor()
        
        assert processor.max_file_size_mb == 100
        assert processor.chunk_size == 10000
        assert hasattr(processor, 'logger')
        assert processor.SUPPORTED_EXTENSIONS == {'.xlsx', '.xls'}
        assert processor.METADATA_PREVIEW_ROWS == 100

    def test_init_custom_parameters(self):
        """Test ExcelProcessor initialization with custom parameters."""
        processor = ExcelProcessor(max_file_size_mb=50, chunk_size=5000)
        
        assert processor.max_file_size_mb == 50
        assert processor.chunk_size == 5000

    def test_supported_extensions_constant(self, excel_processor):
        """Test that supported extensions are properly defined."""
        supported = excel_processor.SUPPORTED_EXTENSIONS
        
        assert isinstance(supported, set)
        assert '.xlsx' in supported
        assert '.xls' in supported
        assert len(supported) == 2

    def test_metadata_preview_rows_constant(self, excel_processor):
        """Test metadata preview rows constant."""
        assert excel_processor.METADATA_PREVIEW_ROWS == 100
        assert isinstance(excel_processor.METADATA_PREVIEW_ROWS, int)
        assert excel_processor.METADATA_PREVIEW_ROWS > 0

    def test_excel_processing_error_inheritance(self):
        """Test ExcelProcessingError is properly defined."""
        error = ExcelProcessingError("Test error")
        assert isinstance(error, Exception)
        assert str(error) == "Test error"


class TestExcelProcessorValidation:
    """Test suite for file validation functionality."""

    @pytest.fixture
    def excel_processor(self):
        """Create a fresh ExcelProcessor instance for testing."""
        return ExcelProcessor()

    @pytest.fixture
    def mock_file_stats(self):
        """Mock file statistics for testing."""
        stats = Mock()
        stats.st_size = 1024 * 1024  # 1MB
        return stats

    @patch('pathlib.Path.exists')
    @patch('pathlib.Path.is_file')
    @patch('pathlib.Path.stat')
    def test_validate_file_success(self, mock_stat, mock_is_file, mock_exists, excel_processor, mock_file_stats):
        """Test successful file validation."""
        mock_exists.return_value = True
        mock_is_file.return_value = True
        mock_stat.return_value = mock_file_stats
        
        test_file = Path("test.xlsx")
        
        # Should not raise exception
        result = excel_processor._validate_file(test_file)
        assert result is None

    @patch('pathlib.Path.exists')
    def test_validate_file_not_found(self, mock_exists, excel_processor):
        """Test file validation when file doesn't exist."""
        mock_exists.return_value = False
        
        test_file = Path("nonexistent.xlsx")
        
        with pytest.raises(ExcelProcessingError) as exc_info:
            excel_processor._validate_file(test_file)
        
        assert "File not found" in str(exc_info.value)

    @patch('pathlib.Path.exists')
    @patch('pathlib.Path.is_file')
    def test_validate_file_not_a_file(self, mock_is_file, mock_exists, excel_processor):
        """Test file validation when path is not a file."""
        mock_exists.return_value = True
        mock_is_file.return_value = False  # It's a directory
        
        test_file = Path("directory.xlsx")
        
        with pytest.raises(ExcelProcessingError) as exc_info:
            excel_processor._validate_file(test_file)
        
        assert "Path is not a file" in str(exc_info.value)

    @patch('pathlib.Path.exists')
    @patch('pathlib.Path.is_file')
    @patch('pathlib.Path.stat')
    def test_validate_file_too_large(self, mock_stat, mock_is_file, mock_exists, excel_processor):
        """Test file validation when file is too large."""
        mock_exists.return_value = True
        mock_is_file.return_value = True
        
        # Create large file stats
        large_stats = Mock()
        large_stats.st_size = 200 * 1024 * 1024  # 200MB (larger than 100MB limit)
        mock_stat.return_value = large_stats
        
        test_file = Path("large.xlsx")
        
        with pytest.raises(ExcelProcessingError) as exc_info:
            excel_processor._validate_file(test_file)
        
        assert "File too large" in str(exc_info.value)

    @patch('pathlib.Path.exists')
    @patch('pathlib.Path.is_file')
    @patch('pathlib.Path.stat')
    def test_validate_file_unsupported_extension(self, mock_stat, mock_is_file, mock_exists, excel_processor, mock_file_stats):
        """Test file validation with unsupported extension."""
        mock_exists.return_value = True
        mock_is_file.return_value = True
        mock_stat.return_value = mock_file_stats
        
        test_file = Path("test.csv")
        
        with pytest.raises(ExcelProcessingError) as exc_info:
            excel_processor._validate_file(test_file)
        
        assert "Unsupported file extension" in str(exc_info.value)


class TestExcelProcessorWorksheetExtraction:
    """Test suite for worksheet extraction functionality."""

    @pytest.fixture
    def excel_processor(self):
        """Create a fresh ExcelProcessor instance for testing."""
        return ExcelProcessor()

    @pytest.fixture
    def mock_workbook_info(self):
        """Mock workbook info for testing."""
        return {
            'Sheet1': {
                'name': 'Sheet1',
                'row_count': 100,
                'column_count': 5,
                'has_data': True
            },
            'Sheet2': {
                'name': 'Sheet2', 
                'row_count': 50,
                'column_count': 3,
                'has_data': True
            }
        }

    @patch('pathlib.Path.exists')
    @patch('pathlib.Path.is_file')
    @patch('pathlib.Path.stat')
    @patch.object(ExcelProcessor, '_get_workbook_info')
    @patch.object(ExcelProcessor, '_process_worksheet')
    def test_extract_worksheets_success(self, mock_process_worksheet, mock_get_workbook_info, 
                                       mock_stat, mock_is_file, mock_exists, excel_processor, mock_workbook_info):
        """Test successful worksheet extraction."""
        # Setup mocks
        mock_exists.return_value = True
        mock_is_file.return_value = True
        mock_stat.return_value.st_size = 1024
        mock_get_workbook_info.return_value = mock_workbook_info
        
        # Mock worksheet data
        mock_worksheet_data = WorksheetData(
            worksheet_name="Sheet1",
            data_frame=pd.DataFrame({'A': [1, 2], 'B': [3, 4]}),
            confidence_score=0.9,
            row_count=2,
            column_count=2,
            has_headers=True
        )
        mock_process_worksheet.return_value = mock_worksheet_data
        
        test_file = Path("test.xlsx")
        
        result = excel_processor._extract_worksheets(test_file)
        
        assert isinstance(result, list)
        assert len(result) == 2  # Two worksheets should be processed
        mock_get_workbook_info.assert_called_once_with(test_file)

    @patch('pathlib.Path.exists')
    @patch('pathlib.Path.is_file') 
    @patch('pathlib.Path.stat')
    @patch.object(ExcelProcessor, '_get_workbook_info')
    def test_extract_worksheets_no_data(self, mock_get_workbook_info, mock_stat, mock_is_file, mock_exists, excel_processor):
        """Test worksheet extraction when no worksheets have data."""
        mock_exists.return_value = True
        mock_is_file.return_value = True
        mock_stat.return_value.st_size = 1024
        
        # Mock workbook with no data worksheets
        mock_get_workbook_info.return_value = {
            'Sheet1': {
                'name': 'Sheet1',
                'row_count': 0,
                'column_count': 0,
                'has_data': False
            }
        }
        
        test_file = Path("empty.xlsx")
        
        result = excel_processor._extract_worksheets(test_file)
        
        assert isinstance(result, list)
        assert len(result) == 0  # No worksheets with data


class TestExcelProcessorWorkbookInfo:
    """Test suite for workbook information extraction."""

    @pytest.fixture
    def excel_processor(self):
        """Create a fresh ExcelProcessor instance for testing."""
        return ExcelProcessor()

    @patch('pathlib.Path.exists')
    @patch('pathlib.Path.is_file')
    @patch('pathlib.Path.stat')
    @patch('openpyxl.load_workbook')
    def test_get_workbook_info_success(self, mock_load_workbook, mock_stat, mock_is_file, mock_exists, excel_processor):
        """Test successful workbook info extraction."""
        mock_exists.return_value = True
        mock_is_file.return_value = True
        mock_stat.return_value.st_size = 1024
        
        # Create mock worksheet
        mock_worksheet = Mock()
        mock_worksheet.title = "TestSheet"
        mock_worksheet.max_row = 10
        mock_worksheet.max_column = 5
        
        # Create mock workbook
        mock_workbook = Mock()
        mock_workbook.worksheets = [mock_worksheet]
        mock_load_workbook.return_value = mock_workbook
        
        test_file = Path("test.xlsx")
        
        result = excel_processor._get_workbook_info(test_file)
        
        assert isinstance(result, dict)
        assert "TestSheet" in result
        assert result["TestSheet"]["name"] == "TestSheet"
        assert result["TestSheet"]["row_count"] == 10
        assert result["TestSheet"]["column_count"] == 5

    @patch('pathlib.Path.exists')
    @patch('pathlib.Path.is_file')
    @patch('pathlib.Path.stat')
    @patch('openpyxl.load_workbook')
    def test_get_workbook_info_invalid_file(self, mock_load_workbook, mock_stat, mock_is_file, mock_exists, excel_processor):
        """Test workbook info extraction with invalid file."""
        mock_exists.return_value = True
        mock_is_file.return_value = True
        mock_stat.return_value.st_size = 1024
        mock_load_workbook.side_effect = InvalidFileException("Not a valid Excel file")
        
        test_file = Path("invalid.xlsx")
        
        with pytest.raises(ExcelProcessingError) as exc_info:
            excel_processor._get_workbook_info(test_file)
        
        assert "Failed to read workbook" in str(exc_info.value)


class TestExcelProcessorDataReading:
    """Test suite for worksheet data reading functionality."""

    @pytest.fixture
    def excel_processor(self):
        """Create a fresh ExcelProcessor instance for testing."""
        return ExcelProcessor()

    @pytest.fixture
    def sample_dataframe(self):
        """Sample DataFrame for testing."""
        return pd.DataFrame({
            'Name': ['Alice', 'Bob', 'Charlie'],
            'Age': [25, 30, 35],
            'City': ['New York', 'London', 'Paris']
        })

    @patch('pandas.ExcelFile')
    def test_read_worksheet_data_success(self, mock_excel_file_class, excel_processor, sample_dataframe):
        """Test successful worksheet data reading."""
        # Create mock ExcelFile instance
        mock_excel_file = Mock()
        mock_excel_file.parse.return_value = sample_dataframe
        mock_excel_file_class.return_value = mock_excel_file
        
        result = excel_processor._read_worksheet_data(mock_excel_file, "Sheet1")
        
        assert isinstance(result, pd.DataFrame)
        assert len(result) == 3
        assert list(result.columns) == ['Name', 'Age', 'City']
        mock_excel_file.parse.assert_called_once_with("Sheet1", header=0)

    @patch('pandas.ExcelFile')
    def test_read_worksheet_data_no_headers(self, mock_excel_file_class, excel_processor, sample_dataframe):
        """Test worksheet data reading without headers."""
        mock_excel_file = Mock()
        mock_excel_file.parse.return_value = sample_dataframe
        mock_excel_file_class.return_value = mock_excel_file
        
        result = excel_processor._read_worksheet_data(mock_excel_file, "Sheet1", has_headers=False)
        
        assert isinstance(result, pd.DataFrame)
        mock_excel_file.parse.assert_called_once_with("Sheet1", header=None)

    @patch('pandas.ExcelFile')
    def test_read_worksheet_data_sheet_not_found(self, mock_excel_file_class, excel_processor):
        """Test worksheet data reading with invalid sheet name."""
        mock_excel_file = Mock()
        mock_excel_file.parse.side_effect = ValueError("Sheet 'NonExistent' not found")
        mock_excel_file_class.return_value = mock_excel_file
        
        with pytest.raises(ExcelProcessingError) as exc_info:
            excel_processor._read_worksheet_data(mock_excel_file, "NonExistent")
        
        assert "Failed to read worksheet data" in str(exc_info.value)


class TestExcelProcessorMetadataEnhancement:
    """Test suite for metadata enhancement functionality."""

    @pytest.fixture
    def excel_processor(self):
        """Create a fresh ExcelProcessor instance for testing.""" 
        return ExcelProcessor()

    @pytest.fixture
    def sample_dataframe(self):
        """Sample DataFrame with mixed data types."""
        return pd.DataFrame({
            'Integer': [1, 2, 3],
            'Float': [1.1, 2.2, 3.3],
            'String': ['a', 'b', 'c'],
            'Boolean': [True, False, True]
        })

    def test_enhance_metadata_basic(self, excel_processor, sample_dataframe):
        """Test basic metadata enhancement."""
        base_metadata = {
            'name': 'TestSheet',
            'row_count': 3,
            'column_count': 4
        }
        
        result = excel_processor._enhance_metadata(sample_dataframe, base_metadata)
        
        assert isinstance(result, dict)
        assert result['name'] == 'TestSheet'
        assert result['row_count'] == 3
        assert result['column_count'] == 4
        assert 'column_types' in result
        assert 'has_headers' in result

    def test_enhance_metadata_empty_dataframe(self, excel_processor):
        """Test metadata enhancement with empty DataFrame."""
        empty_df = pd.DataFrame()
        base_metadata = {
            'name': 'EmptySheet',
            'row_count': 0,
            'column_count': 0
        }
        
        result = excel_processor._enhance_metadata(empty_df, base_metadata)
        
        assert isinstance(result, dict)
        assert result['name'] == 'EmptySheet'
        assert result['row_count'] == 0
        assert result['column_count'] == 0


class TestExcelProcessorEndToEnd:
    """Test suite for end-to-end Excel processing functionality."""

    @pytest.fixture
    def excel_processor(self):
        """Create a fresh ExcelProcessor instance for testing."""
        return ExcelProcessor()

    @patch('pathlib.Path.exists')
    @patch('pathlib.Path.is_file')
    @patch('pathlib.Path.stat')
    @patch.object(ExcelProcessor, '_extract_worksheets')
    def test_process_file_success(self, mock_extract_worksheets, mock_stat, mock_is_file, mock_exists, excel_processor):
        """Test successful end-to-end file processing."""
        # Setup mocks
        mock_exists.return_value = True
        mock_is_file.return_value = True
        mock_stat.return_value.st_size = 1024
        
        # Mock worksheet data
        mock_worksheets = [
            WorksheetData(
                worksheet_name="Sheet1",
                data_frame=pd.DataFrame({'A': [1, 2], 'B': [3, 4]}),
                confidence_score=0.9,
                row_count=2,
                column_count=2,
                has_headers=True
            )
        ]
        mock_extract_worksheets.return_value = mock_worksheets
        
        test_file = Path("test.xlsx")
        
        result = excel_processor.process_file(test_file)
        
        assert isinstance(result, list)
        assert len(result) == 1
        assert result[0].worksheet_name == "Sheet1"
        assert result[0].row_count == 2

    @patch('pathlib.Path.exists')
    def test_process_file_validation_failure(self, mock_exists, excel_processor):
        """Test file processing with validation failure."""
        mock_exists.return_value = False
        
        test_file = Path("nonexistent.xlsx")
        
        with pytest.raises(ExcelProcessingError) as exc_info:
            excel_processor.process_file(test_file)
        
        assert "Excel processing failed" in str(exc_info.value)

    @patch('pathlib.Path.exists')
    @patch('pathlib.Path.is_file')
    @patch('pathlib.Path.stat')
    @patch.object(ExcelProcessor, '_extract_worksheets')
    def test_process_file_extraction_failure(self, mock_extract_worksheets, mock_stat, mock_is_file, mock_exists, excel_processor):
        """Test file processing with extraction failure."""
        mock_exists.return_value = True
        mock_is_file.return_value = True
        mock_stat.return_value.st_size = 1024
        mock_extract_worksheets.side_effect = Exception("Extraction failed")
        
        test_file = Path("test.xlsx")
        
        with pytest.raises(ExcelProcessingError) as exc_info:
            excel_processor.process_file(test_file)
        
        assert "Excel processing failed" in str(exc_info.value)